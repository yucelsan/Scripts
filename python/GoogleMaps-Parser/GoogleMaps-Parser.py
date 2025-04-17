from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.common.keys import Keys
from selenium.webdriver.chrome.service import Service
from webdriver_manager.chrome import ChromeDriverManager
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
import time
import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill
from openpyxl.utils import get_column_letter
import re

def extract_phone_from_page():
    try:
        all_elements = driver.find_elements(By.XPATH, "//button | //span | //div")
        for el in all_elements:
            try:
                text = el.get_attribute("aria-label") or el.text
                if text and re.search(r"0[1-9](?:[\s.-]?\d{2}){4}", text):
                    return re.search(r"0[1-9](?:[\s.-]?\d{2}){4}", text).group()
            except:
                continue
    except:
        pass
    return ""

# Config Chrome
options = webdriver.ChromeOptions()
options.add_argument("--start-maximized")
driver = webdriver.Chrome(service=Service(ChromeDriverManager().install()), options=options)

# Lancer Google Maps
driver.get("https://www.google.com/maps")
wait = WebDriverWait(driver, 10)
search_box = wait.until(EC.presence_of_element_located((By.ID, "searchboxinput")))
search_box.send_keys("entreprise batiment grenoble")
search_box.send_keys(Keys.ENTER)
time.sleep(5)

# Scroll les résultats pour en charger plus
scrollable_div = driver.find_element(By.XPATH, '//div[contains(@class, "m6QErb")]')
for _ in range(20):
    driver.execute_script("arguments[0].scrollTop = arguments[0].scrollHeight", scrollable_div)
    time.sleep(1.5)

# Extraction
results = driver.find_elements(By.CLASS_NAME, "Nv2PK")
data = []

for index, res in enumerate(results):
    try:
        print(f"Fiche {index + 1}/{len(results)}")
        driver.execute_script("arguments[0].scrollIntoView(true);", res)
        res.click()
        time.sleep(2.5)

        wait = WebDriverWait(driver, 5)
        wait.until(EC.presence_of_element_located((By.XPATH, "//h1[contains(@class,'DUwDvf')]")))

        try:
            name = driver.find_element(By.XPATH, "//h1[contains(@class,'DUwDvf')]").text
        except:
            name = ""

        try:
            address = driver.find_element(By.XPATH, "//button[contains(@aria-label, 'Adresse')]").text
        except:
            address = ""

        phone = extract_phone_from_page()

        try:
            rating = driver.find_element(By.CLASS_NAME, "MW4etd").text
        except:
            rating = ""

        try:
            review_count = driver.find_element(By.CLASS_NAME, "UY7F9").text
        except:
            review_count = ""

        try:
            site_elem = driver.find_element(By.XPATH, "//a[contains(@aria-label, 'Site') and starts-with(@href, 'http')]")
            site = site_elem.get_attribute("href")
        except:
            site = "Pas de site"

        data.append({
            "Nom": name,
            "Adresse": address,
            "Note": rating,
            "Avis": review_count,
            "Site Web": site,
            "Téléphone": phone
        })

    except Exception as e:
        print(f"Erreur fiche {index + 1} : {e}")
        continue

# Filtrer uniquement ceux sans site web
# data = [entry for entry in data if entry["Site Web"] == "Pas de site"]

# Excel export stylisé
df = pd.DataFrame(data)
excel_path = "entreprises_batiment_grenoble.xlsx"
df.to_excel(excel_path, index=False)

wb = load_workbook(excel_path)
ws = wb.active
ws.freeze_panes = ws["A2"]
ws.auto_filter.ref = ws.dimensions

header_fill = PatternFill(start_color="4F81BD", end_color="4F81BD", fill_type="solid")
header_font = Font(color="FFFFFF", bold=True)

for cell in ws[1]:
    cell.fill = header_fill
    cell.font = header_font

for col in ws.columns:
    max_length = 0
    col_letter = get_column_letter(col[0].column)
    for cell in col:
        if cell.value:
            max_length = max(max_length, len(str(cell.value)))
    ws.column_dimensions[col_letter].width = max_length + 2

wb.save(excel_path)
print("Fichier Excel propre généré :", excel_path)

driver.quit()
